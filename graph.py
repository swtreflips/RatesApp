"""
graph.py — Outbound notification reference implementation (Send Rate Request / Send Reminder)

Companion to ALERTS.md. This is the *actual* working code for the outbound email action,
reusing the delegated Microsoft Graph setup from the internal CRM (auth.py / send.ipynb).

WHAT THIS IS
------------
A runnable Python reference + local test harness for the send action described in
ALERTS.md §6 / §6a / §6c. Two things in here are used for real:
  1. `seed_refresh_token()` — the one-time device-code login that produces the refresh token
     you store in Supabase (`graph_credentials` row). RUN THIS LOCALLY.
  2. the refresh-grant + `/me/sendMail` + template-fill logic — the exact mechanics the
     Supabase Edge Function (`notify-forwarders`, Deno/TS) will mirror. You can validate the
     whole chain here (real email, real .xlsx attachment) BEFORE building the function.

It deliberately does NOT talk to Supabase. The "which forwarders / which lanes" decision is
DB-side (ALERTS.md §3/§14 — outstanding-per-forwarder). Here you pass that data in as
`recipients`, matching the §6 per-forwarder payload shape:
    recipients = [
        { "name": "Tanera Transport", "email": "silvia@example.com",
          "lanes": [ {"pol": "...", "fd": "...", "pod": None, "last_cy": None,
                      "container_type": "40' HC", "container_count": 3}, ... ] },
        ...
    ]

DEPENDENCIES
------------
    pip install msal requests openpyxl

REUSED VARIABLES (from the CRM auth.py — delegated public client; NO client secret)
-----------------------------------------------------------------------------------
    client_id  = d4a32e7f-44b5-472b-9a4b-bdfd0c43a61f
    tenant_id  = cc38fec1-5ec4-40bd-985a-4855a2ba4372
    sender     = luismht@primetimepackaging.com   (Phase 1; Phase 2 → shared rates@ mailbox)

CLI
---
    python graph.py seed                       # one-time: device-code login → save refresh token
    python graph.py send-template [to]         # BASIC: email the raw template (default: to luismht@)
    python graph.py refresh                    # keep-alive: redeem + rotate the token (no send)
    python graph.py test-send you@example.com  # send yourself a sample (filled) rate-request email
    python graph.py test-reminder you@example.com
"""

from __future__ import annotations

import base64
import io
import json
import os
import sys
from datetime import date

import requests

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

CLIENT_ID = "d4a32e7f-44b5-472b-9a4b-bdfd0c43a61f"
TENANT_ID = "cc38fec1-5ec4-40bd-985a-4855a2ba4372"
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
TOKEN_ENDPOINT = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"

# Phase 1 sender = the seeded mailbox. Sends go through /me, so this is just informational
# (the mailbox is whichever account completed the device-code login in seed_refresh_token()).
SENDER = "luismht@primetimepackaging.com"

# Delegated scopes. MSAL auto-adds offline_access/openid/profile during seeding; the raw
# refresh grant requests them explicitly.
SCOPES = ["Mail.Send"]
REFRESH_SCOPE = "https://graph.microsoft.com/Mail.Send offline_access"

GRAPH_SENDMAIL = "https://graph.microsoft.com/v1.0/me/sendMail"

APP_URL = "https://rates.ptpbags.com"          # link in the invitation body
SENDER_NAME = "Luis"                            # signature in the invitation body

# Template lives at the repo root for now (ALERTS.md §6a says move it to Supabase Storage later).
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "PTP OFQ Rates Template.xlsx")

# Local refresh-token store for this reference. In production this is the locked
# `graph_credentials` Supabase row (ALERTS.md §6c), read+rewritten by the Edge Function.
REFRESH_TOKEN_FILE = os.path.join(os.path.dirname(__file__), ".graph_refresh_token")

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# Template column map (1-based), Sheet1. See ALERTS.md §6a. Forwarder columns (J–S) stay blank;
# A,B,E,F,G are deferred request fields left blank. Data rows start at row 2.
COL_POL = 3             # C  Port of Loading
COL_FD = 4              # D  Final Destination
COL_CONTAINER_TYPE = 8  # H  Container Type
COL_CONTAINER_COUNT = 9 # I  # of Containers
COL_FORWARDER = 12      # L  Forwarder (pre-filled with recipient name)
COL_POD = 13            # M  Port of Discharge (only if requester set it)
COL_LAST_CY = 14        # N  Last CY/CFS (only if requester set it)
DATA_START_ROW = 2


# ─────────────────────────────────────────────────────────────────────────────
# REFRESH-TOKEN STORE (swap these two for the `graph_credentials` row in production)
# ─────────────────────────────────────────────────────────────────────────────

def load_refresh_token() -> str:
    if not os.path.exists(REFRESH_TOKEN_FILE):
        raise RuntimeError(
            "No refresh token found. Run `python graph.py seed` once to create it."
        )
    with open(REFRESH_TOKEN_FILE, "r", encoding="utf-8") as f:
        return f.read().strip()


def save_refresh_token(token: str) -> None:
    with open(REFRESH_TOKEN_FILE, "w", encoding="utf-8") as f:
        f.write(token)


# ─────────────────────────────────────────────────────────────────────────────
# AUTH
# ─────────────────────────────────────────────────────────────────────────────

def seed_refresh_token() -> str:
    """
    One-time, interactive. Reuses the CRM's device-code flow to obtain a refresh token and
    persist it. Run locally; copy the resulting token into the Supabase `graph_credentials`
    row for the Edge Function.
    """
    import msal  # imported here so the send path doesn't require msal

    # A SerializableTokenCache is required to read the refresh token back out — the default
    # in-memory TokenCache has no .serialize().
    cache = msal.SerializableTokenCache()
    app = msal.PublicClientApplication(CLIENT_ID, authority=AUTHORITY, token_cache=cache)

    flow = app.initiate_device_flow(scopes=SCOPES)
    if "user_code" not in flow:
        raise RuntimeError(f"Failed to start device flow: {json.dumps(flow, indent=2)}")
    print(flow["message"])  # "To sign in, use a web browser to open … and enter the code …"

    result = app.acquire_token_by_device_flow(flow)
    if "access_token" not in result:
        raise RuntimeError(
            "Auth failed: " + str(result.get("error_description", result))
        )

    # The refresh token lives in the serialized MSAL cache, not the result dict.
    cache_json = json.loads(cache.serialize())
    rt_entries = cache_json.get("RefreshToken", {})
    if not rt_entries:
        raise RuntimeError("No refresh token in cache — ensure offline_access was granted.")
    refresh_token = next(iter(rt_entries.values()))["secret"]

    save_refresh_token(refresh_token)
    print(f"\nRefresh token saved to {REFRESH_TOKEN_FILE}")
    print("→ In production, store this value in the Supabase `graph_credentials` row.")
    return refresh_token


def get_access_token() -> str:
    """
    Redeem the stored refresh token for an access token (public-client refresh grant — no
    client secret), and PERSIST the rotated refresh token. This is exactly what the Edge
    Function does per send (ALERTS.md §6c).
    """
    refresh_token = load_refresh_token()
    resp = requests.post(
        TOKEN_ENDPOINT,
        data={
            "grant_type": "refresh_token",
            "client_id": CLIENT_ID,
            "refresh_token": refresh_token,
            "scope": REFRESH_SCOPE,
        },
        timeout=30,
    )
    if resp.status_code != 200:
        # ~90-day lapse, MFA, or a conditional-access change invalidates the RT → re-seed.
        raise RuntimeError(
            "Refresh-token grant failed (re-run `python graph.py seed`): " + resp.text
        )
    data = resp.json()
    # Entra rotates the refresh token on each use — persist the new one or the chain dies.
    if data.get("refresh_token"):
        save_refresh_token(data["refresh_token"])
    return data["access_token"]


# ─────────────────────────────────────────────────────────────────────────────
# EMAIL COMPOSITION (ALERTS.md §6a)
# ─────────────────────────────────────────────────────────────────────────────

def build_invitation_html(forwarder_name: str, *, is_reminder: bool) -> str:
    """The short invitation body. Lanes are NOT listed here — they ride in the attachment."""
    lead = (
        "Just following up — we still need your rates for the lanes in the attached sheet."
        if is_reminder
        else "Attached is the rate request with the lanes we need quoted for this period."
    )
    return f"""\
<p>Hello {forwarder_name},</p>
<p>{lead} Please complete the sheet (rate, carrier, port of discharge, last CY, free days…)
and indicate whether each shipment is <b>direct or involves a transshipment</b>.</p>
<p>Return your rates by uploading the completed sheet in our portal, or entering them directly:
<a href="{APP_URL}">{APP_URL}</a></p>
<p>Thanks,<br>{SENDER_NAME}</p>
"""


def fill_template(lanes: list[dict], forwarder_name: str) -> bytes:
    """
    Load PTP OFQ Rates Template.xlsx, write the requester columns for each lane (forwarder
    columns left blank), and return the .xlsx bytes. openpyxl preserves the workbook's other
    sheet (Validation) and most formatting/data-validation; verify dropdowns survive in Excel.
    """
    from openpyxl import load_workbook  # imported here so auth/send don't require openpyxl

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Sheet1"]

    r = DATA_START_ROW
    for lane in lanes:
        ws.cell(row=r, column=COL_POL, value=lane.get("pol"))
        ws.cell(row=r, column=COL_FD, value=lane.get("fd"))
        ws.cell(row=r, column=COL_CONTAINER_TYPE, value=lane.get("container_type"))
        ws.cell(row=r, column=COL_CONTAINER_COUNT, value=lane.get("container_count"))
        ws.cell(row=r, column=COL_FORWARDER, value=forwarder_name)
        # POD / Last CY only if the requester pre-set them (optional refinement); else leave
        # blank for the forwarder to fill.
        if lane.get("pod"):
            ws.cell(row=r, column=COL_POD, value=lane["pod"])
        if lane.get("last_cy"):
            ws.cell(row=r, column=COL_LAST_CY, value=lane["last_cy"])
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def attachment_name(forwarder_name: str) -> str:
    safe = forwarder_name.strip().replace("/", "-")
    return f"PTP OFQ Rates - {safe} - {date.today().isoformat()}.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# SEND (ALERTS.md §6c) — /me/sendMail with HTML body + .xlsx attachment
# ─────────────────────────────────────────────────────────────────────────────

def send_mail(
    access_token: str,
    to_email: str,
    subject: str,
    html_body: str,
    attachment_bytes: bytes,
    attachment_filename: str,
) -> None:
    """POST /v1.0/me/sendMail. Sends as the seeded mailbox. Raises on non-202."""
    payload = {
        "message": {
            "subject": subject,
            "body": {"contentType": "HTML", "content": html_body},
            "toRecipients": [{"emailAddress": {"address": to_email}}],
            "attachments": [
                {
                    "@odata.type": "#microsoft.graph.fileAttachment",
                    "name": attachment_filename,
                    "contentType": XLSX_CONTENT_TYPE,
                    "contentBytes": base64.b64encode(attachment_bytes).decode("ascii"),
                }
            ],
        },
        "saveToSentItems": True,
    }
    resp = requests.post(
        GRAPH_SENDMAIL,
        headers={
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=60,
    )
    if resp.status_code != 202:
        raise RuntimeError(f"sendMail failed for {to_email}: {resp.status_code} {resp.text}")


# ─────────────────────────────────────────────────────────────────────────────
# BASIC SEND + KEEP-ALIVE (Phase-1 validation — the smallest end-to-end pieces)
# ─────────────────────────────────────────────────────────────────────────────

def send_template(to_email: str = SENDER) -> None:
    """
    Smallest real send: attach the RAW PTP OFQ Rates Template.xlsx (no fill) and email it.
    Proves the whole chain — stored refresh token → access token (rotated) → /me/sendMail with
    an .xlsx attachment. Defaults to the seeded mailbox (sends to yourself).
    """
    with open(TEMPLATE_PATH, "rb") as f:
        attachment = f.read()

    token = get_access_token()  # refresh-grant + persists the rotated token (keep-alive on use)
    body = (
        "<p>Hello,</p>"
        "<p>Attached is the rate request template "
        "(<b>PTP OFQ Rates Template.xlsx</b>).</p>"
        f"<p>Thanks,<br>{SENDER_NAME}</p>"
    )
    send_mail(token, to_email, "PTP OFQ Rates Template", body,
              attachment, os.path.basename(TEMPLATE_PATH))
    print(f"✅ sent the template to {to_email}")


def refresh_token_keepalive() -> None:
    """
    Keep-alive: redeem + ROTATE the refresh token without sending anything. This is exactly the
    call the production Supabase pg_cron job runs every 7 days (ALERTS.md §6c). Schedule it (e.g.
    weekly) so the token never lapses during quiet stretches; each real send keeps it fresh too.
    """
    get_access_token()
    print("✅ refresh token redeemed + rotated (chain alive)")


# ─────────────────────────────────────────────────────────────────────────────
# ACTIONS — Send Rate Request / Send Reminder (one engine, ALERTS.md §3/§14)
# ─────────────────────────────────────────────────────────────────────────────

def _send_to_recipients(recipients: list[dict], *, is_reminder: bool, period: int | None) -> dict:
    """
    Loop the recipients, fill each one's sheet from THEIR lanes, and send. `recipients` is
    pre-computed DB-side (outstanding-per-forwarder); recipients with no lanes are skipped.
    Returns {"sent": [...], "failed": [...]}. The Edge Function mirrors this loop and also
    writes the `notifications` / `notification_recipients` audit rows (ALERTS.md §7).
    """
    label = "Reminder: Rate Request" if is_reminder else "Rate Request"
    subject = f"{label} — Period {period}" if period is not None else label

    token = get_access_token()
    sent, failed = [], []

    for rcpt in recipients:
        lanes = rcpt.get("lanes") or []
        if not lanes:
            continue  # nothing outstanding for this forwarder → don't email
        try:
            xlsx = fill_template(lanes, rcpt["name"])
            send_mail(
                token,
                rcpt["email"],
                subject,
                build_invitation_html(rcpt["name"], is_reminder=is_reminder),
                xlsx,
                attachment_name(rcpt["name"]),
            )
            sent.append(rcpt["email"])
            print(f"  ✅ sent to {rcpt['name']} <{rcpt['email']}> ({len(lanes)} lane(s))")
        except Exception as exc:  # noqa: BLE001 — record per-recipient and continue
            failed.append({"email": rcpt["email"], "error": str(exc)})
            print(f"  ❌ {rcpt['name']} <{rcpt['email']}>: {exc}")

    return {"sent": sent, "failed": failed}


def send_rate_request(recipients: list[dict], period: int | None = None) -> dict:
    """Initial / subsequent request blast. `recipients` = selected forwarders + their full
    outstanding lanes."""
    return _send_to_recipients(recipients, is_reminder=False, period=period)


def send_reminder(recipients: list[dict], period: int | None = None) -> dict:
    """Reminder. Same engine; callers pass only non-responding forwarders + their outstanding
    lanes (ALERTS.md §3) and the copy/subject shift to a nudge."""
    return _send_to_recipients(recipients, is_reminder=True, period=period)


# ─────────────────────────────────────────────────────────────────────────────
# CLI / local test harness
# ─────────────────────────────────────────────────────────────────────────────

def _sample_recipient(email: str) -> dict:
    return {
        "name": "Test Forwarder",
        "email": email,
        "lanes": [
            {"pol": "Nhava Sheva, India", "fd": "Commerce, CA", "pod": None,
             "last_cy": None, "container_type": "40' HC", "container_count": 3},
            {"pol": "Shanghai, China", "fd": "Chicago, IL", "pod": "Los Angeles, CA",
             "last_cy": None, "container_type": "40' GP", "container_count": 1},
        ],
    }


def _main(argv: list[str]) -> int:
    if not argv:
        print(__doc__)
        return 0
    cmd = argv[0]

    if cmd == "seed":
        seed_refresh_token()
        return 0

    if cmd == "send-template":
        send_template(argv[1] if len(argv) > 1 else SENDER)
        return 0

    if cmd == "refresh":
        refresh_token_keepalive()
        return 0

    if cmd in ("test-send", "test-reminder"):
        if len(argv) < 2:
            print(f"usage: python graph.py {cmd} <to_email>")
            return 2
        recipients = [_sample_recipient(argv[1])]
        result = (send_reminder if cmd == "test-reminder" else send_rate_request)(
            recipients, period=1
        )
        print(json.dumps(result, indent=2))
        return 0 if not result["failed"] else 1

    print(f"unknown command: {cmd}")
    print(__doc__)
    return 2


if __name__ == "__main__":
    raise SystemExit(_main(sys.argv[1:]))
